from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
from marshmallow import Schema, fields

app = Flask(__name__)
wb = load_workbook("frameworks.xlsx")
ws = wb.active

class Framework:
    def __init__(self, id, name):
        self.id = id
        self.name = name

    def __repr__(self):

        return "<Framework %s>" % self.name

class FrameworkSchema(Schema):
    id = fields.Int()
    name = fields.Str()

# Defining marshmallow schemas
framework_schema = FrameworkSchema()
frameworks_schema = FrameworkSchema(many=True)

@app.route("/")
def index():
    """
    This route is irrelevant for the REST API, you can delete it if you want to.
    """

    return "Hello World!"

@app.route("/api/frameworks/")
def get_frameworks():
    framework_objects = []
    frameworks = list(ws.rows)

    for f in frameworks:
        framework_objects.append(
                Framework(f[0].value, f[1].value)
                )

    result, errors = frameworks_schema.dump(framework_objects)

    return jsonify(result)

@app.route("/api/frameworks/<string:name>")
def get_framework_by_name(name):
    r = ws.max_row
    c = ws.max_column

    for i in range(1, r+1):
        if ws.cell(row=i, column=2).value == name:
            framework_row = i

    framework = Framework(
                id=ws.cell(row=framework_row, column=1).value,
                name=ws.cell(row=framework_row, column=2).value
            )
    data, errors = framework_schema.dump(framework)

    return jsonify(data)

@app.route("/api/frameworks/", methods=["POST"])
def add_framework():
    new_framework = {
            "id": request.json["id"],
            "name": request.json["name"]
            }
    ws.append([request.json["id"], request.json["name"]])
    wb.save("frameworks.xlsx")

    return jsonify(new_framework)

@app.route("/api/frameworks/<string:id>", methods=["PUT"])
def edit_framework(id):
    r = ws.max_row
    c = ws.max_column

    for i in range(1, r+1):
        if ws.cell(row=i, column=1).value == id:
            framework_row = i

    ws.cell(row=framework_row, column=1).value = request.json["id"]
    ws.cell(row=framework_row, column=2).value = request.json["name"]
    wb.save("frameworks.xlsx")

    framework = Framework(
                id=ws.cell(row=framework_row, column=1).value,
                name=ws.cell(row=framework_row, column=2).value
            )
    data, errors = framework_schema.dump(framework)

    return jsonify(data)

@app.route("/api/frameworks/<string:id>", methods=["DELETE"])
def delete_framework(id):
    r = ws.max_row
    c = ws.max_column

    for i in range(1, r+1):
        if ws.cell(row=i, column=1).value == id:
            framework_row = i

    ws.cell(row=framework_row, column=1).value = None
    ws.cell(row=framework_row, column=2).value = None
    wb.save("frameworks.xlsx")

    return jsonify({"message": "ok"})


"""
Since Flask 1.0 we are invited to use the following method to execute the app:

GNU/Linux:
    export FLASK_APP=app.py

Windows OS:
    CMD:
        set FLASK_APP=app.py
    Powershell:
        $env:FLASK_APP = "app.py"

Additionally you can set FLASK_ENV=development for Debug Mode

You cand find more info about it here:
http://flask.pocoo.org/docs/1.0/quickstart/#a-minimal-application
"""
